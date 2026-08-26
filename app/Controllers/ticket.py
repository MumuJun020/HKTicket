from flask import Blueprint, jsonify, request, Response, stream_with_context
from playwright.async_api import async_playwright
from ..Auto.ticket_operation import run as run_ticket
from ..Auto.ticket_login import login_all, check_login_status
from ..Auto.ticket_parser import parse_event
from ..Auto import ticket_store as store
from ..Auto.bit_api import getAllBrowsers, BitBrowserNotRunning, clearLoginStateBatch
from ..utils.logger import LogCapture, log_manager, task_state_manager
import uuid
import threading
import asyncio
import json
import queue
import time
from datetime import datetime

ticket = Blueprint("ticket", __name__)


@ticket.route("/do_grab", methods=["post"])
def do_grab():
    """
    发起抢票：按第二步保存的配置，每个账号带各自的场次/票档/张数并发开抢。

    请求体：
    {
        "account_ids": ["..."],       # 参与的抢票人，不传则全部有配置的账号
        "start_at": "2026-08-20 13:00:00",  # 定时抢票；不传或留空 = 立即抢票
        "max_attempts": 0,            # 最大尝试次数，0=不限制
        # 重试间隔别调太小：站点按 IP 限流，实测密集请求几十次就会被掐连接，
        # 被限之后该窗口连正常浏览都不行，反而彻底抢不了
        "retry_interval_min": 2.0,
        "retry_interval_max": 4.0,
        "stop_others_on_success": false  # 多账号各抢各的，默认不因一人成功而停掉其他人
    }
    """
    task_id = str(uuid.uuid4())
    try:
        body = request.json or {}

        accounts = store.get_accounts()
        only_ids = body.get("account_ids")
        if only_ids:
            accounts = [a for a in accounts if a["id"] in only_ids]
        if not accounts:
            return jsonify(code=400, result="没有可参与抢票的抢票人"), 400

        plans = store.get_plans()
        event = store.get_event()

        if not event.get("sessions"):
            return jsonify(code=400, result="还没有解析过活动，请先完成第一步"), 400

        # 当前活动里合法的场次/票档文本，用来校验配置有没有跟活动对上。
        # 这个校验很重要：解析过 A 活动、配好方案，之后又解析了 B 活动，
        # event.json 会被 B 覆盖，但 plans 里还留着 A 的场次文本。
        # 没有校验的话，抢票会导航到 B 的页面去找 A 的场次，
        # 表现成"未找到匹配场次"，看着像选择器坏了，其实是配置串台了。
        valid_sessions = {s["text"] for s in event["sessions"]}
        valid_tiers = {t["text"] for s in event["sessions"] for t in s["tiers"]}
        member_tiers = {t["text"] for s in event["sessions"]
                        for t in s["tiers"] if t.get("need_member_code")}

        # 一个窗口只能有一个抢票人，否则并发抢票时多个任务会操作同一个浏览器
        used_windows = {}
        for a in accounts:
            bid = a.get("browser_id")
            if not bid:
                continue
            label = a.get("remark") or a.get("email")
            if bid in used_windows:
                return (
                    jsonify(
                        code=400,
                        result=f"「{used_windows[bid]}」和「{label}」绑定了同一个浏览器窗口，"
                               "并发抢票时会互相干扰。请先给他们分配各自独立的窗口。",
                    ),
                    400,
                )
            used_windows[bid] = label

        # 窗口 ID -> 界面上显示的序号。战果记录里给的是序号，
        # 因为付款时人要照着它去比特浏览器里找窗口，32 位的 ID 没法用。
        # 拿不到就算了（比特浏览器没开等），序号只是方便人看，
        # 不该因为它让整个抢票起不来。
        seq_by_id = {}
        try:
            seq_by_id = {b["id"]: b.get("seq") for b in getAllBrowsers()}
        except Exception:
            pass

        # 把账号 + 配置组装成引擎要的任务列表，顺便校验配置是否完整。
        # problems 是**导致该账号被跳过**的原因，notices 只是提醒（照常参与抢票）——
        # 两者混在一起的话，日志会把"提醒"也写成"已跳过"，误导得很厉害。
        tasks = []
        problems = []
        notices = []
        for a in accounts:
            label = a.get("remark") or a.get("email")
            if not a.get("browser_id"):
                problems.append(f"{label} 没绑定浏览器窗口")
                continue
            plan = plans.get(a["id"]) or {}
            if not plan.get("session_text") or not plan.get("tier_text"):
                problems.append(f"{label} 还没配置场次或票档")
                continue
            if plan["session_text"] not in valid_sessions:
                problems.append(
                    f"{label} 配的场次「{plan['session_text']}」不属于当前活动"
                    f"（{event.get('name')}），请重新配置"
                )
                continue
            if plan["tier_text"] not in valid_tiers:
                problems.append(
                    f"{label} 配的票档「{plan['tier_text']}」不属于当前活动，请重新配置"
                )
                continue
            # 需会员码的票档**照常参与抢票**，只是提醒一句。
            # 不能在这里拦掉：会员预售场的全部票档都需要码，一拦就等于整场都抢不了。
            # 有码就在弹窗出现时自动填，没码到那一步再停，那是运行时的事。
            if plan["tier_text"] in member_tiers and not (a.get("member_code") or "").strip():
                notices.append(
                    f"{label} 选的票档可能要求会员优先购票码，但他没有填；"
                    f"届时若弹出输入框将无法继续"
                )
            tasks.append(
                {
                    "browser_id": a["browser_id"],
                    "label": label,
                    # 战果记录要用：付款时得知道去哪个号、哪个窗口里操作
                    "email": a.get("email", ""),
                    "member_code": a.get("member_code", ""),
                    "browser_seq": seq_by_id.get(a["browser_id"]),
                    "session_text": plan["session_text"],
                    "tier_text": plan["tier_text"],
                    "quantity": int(plan.get("quantity") or 1),
                }
            )

        if not tasks:
            return jsonify(code=400, result="；".join(problems) or "没有可执行的配置"), 400

        # 定时抢票：把 "YYYY-MM-DD HH:MM:SS" 解析成时间戳
        start_at = None
        start_at_raw = (body.get("start_at") or "").strip()
        if start_at_raw:
            try:
                dt = datetime.strptime(start_at_raw, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                try:
                    # 兼容 datetime-local 控件的 "YYYY-MM-DDTHH:MM" 格式
                    dt = datetime.strptime(start_at_raw, "%Y-%m-%dT%H:%M")
                except ValueError:
                    return (
                        jsonify(code=400, result="开抢时间格式不对，应为 YYYY-MM-DD HH:MM:SS"),
                        400,
                    )
            start_at = dt.timestamp()
            if start_at <= time.time():
                return jsonify(code=400, result="开抢时间已经过了，请重新设置"), 400

        common = {
            "event_url": event.get("event_url", ""),
            "event_url_fragment": body.get("event_url_fragment", "hkticketing"),
            "max_attempts": int(body.get("max_attempts", 0)),
            "retry_interval_min": float(body.get("retry_interval_min", 2.0)),
            "retry_interval_max": float(body.get("retry_interval_max", 4.0)),
            # 抢到后是否自动勾选条款并点「确认订单」完成锁单（仍然不付款）
            "auto_confirm": bool(body.get("auto_confirm", False)),
            # 本轮战果归到哪个批次下。今天抢 A 活动、明天抢 B 活动，
            # 记录不该混在一起，导出时也要能只导某一批。
            "batch_id": store.start_batch(event),
        }
        stop_others = bool(body.get("stop_others_on_success", False))

        _run_in_background(
            task_id,
            lambda pw: run_ticket(
                pw,
                tasks,
                task_id,
                common,
                stop_others_on_success=stop_others,
                start_at=start_at,
            ),
            "抢票任务失败",
            kind="grab",
        )

        if problems:
            log_manager.add_log(task_id, "已跳过：" + "；".join(problems), "warning")
        if notices:
            log_manager.add_log(task_id, "提醒：" + "；".join(notices), "warning")
        mode = f"定时抢票（{start_at_raw}）" if start_at else "立即抢票"
        log_manager.add_log(task_id, f"{mode}，共 {len(tasks)} 个账号参与...", "info")

        return jsonify(
            code=200,
            result=f"{mode}任务已启动，共 {len(tasks)} 个账号",
            task_id=task_id,
        )
    except Exception as e:
        log_manager.add_log(task_id, f"启动抢票任务失败：{str(e)}", "error")
        return (
            jsonify(code=500, result="启动抢票任务失败：" + str(e), task_id=task_id),
            500,
        )


def _run_in_background(task_id, coro_factory, fail_msg, kind=""):
    """
    统一的后台任务启动器：建日志/状态任务 -> 起线程 -> 线程里跑 asyncio 事件循环。

    coro_factory 接收一个 playwright 实例，返回要执行的协程。

    :param kind: 任务类型（grab / login）。刷新页面后要恢复的只有抢票任务。
    """
    log_manager.create_task(task_id)
    task_state_manager.create_task(task_id, kind)

    def _thread():
        try:
            with LogCapture(task_id):

                async def _run():
                    async with async_playwright() as playwright:
                        await coro_factory(playwright)

                asyncio.run(_run())
        except Exception as e:
            log_manager.add_log(task_id, f"{fail_msg}：{str(e)}", "error")
        finally:
            # 必须在 finally 里标结束：异常退出时也要标，否则那个任务会永远
            # 显示成"进行中"，刷新页面后还会被恢复出来，用户点停止也停不掉
            # （线程早没了），只能重启程序。
            task_state_manager.finish_task(task_id)

    threading.Thread(target=_thread, daemon=True).start()


# ------------------------------------------------------------------
# 抢票人管理（增删改查）
# ------------------------------------------------------------------

@ticket.route("/accounts", methods=["get"])
def list_accounts():
    """列出所有抢票人。密码脱敏，不下发到前端。"""
    try:
        return jsonify(code=200, accounts=store.get_accounts_safe())
    except Exception as e:
        return jsonify(code=500, result=str(e)), 500


@ticket.route("/accounts/save", methods=["post"])
def save_account():
    """
    新增或更新抢票人。带 id 为更新，不带为新增。
    更新时密码留空或传 ****** 表示不改密码。
    """
    try:
        saved = store.save_account(request.json or {})
        return jsonify(code=200, result="已保存", id=saved["id"])
    except ValueError as e:
        return jsonify(code=400, result=str(e)), 400
    except Exception as e:
        return jsonify(code=500, result=str(e)), 500


@ticket.route("/accounts/delete_batch", methods=["post"])
def delete_accounts_batch():
    """批量删除抢票人。一次写盘，不会出现删一半的中间状态。"""
    try:
        ids = (request.json or {}).get("ids") or []
        if not ids:
            return jsonify(code=400, result="没有选中任何抢票人"), 400
        n = store.delete_accounts(ids)
        return jsonify(code=200, result=f"已删除 {n} 个抢票人")
    except Exception as e:
        return jsonify(code=500, result=str(e)), 500


@ticket.route("/accounts/delete", methods=["post"])
def delete_account():
    try:
        acc_id = (request.json or {}).get("id")
        if not acc_id:
            return jsonify(code=400, result="缺少 id 参数"), 400
        if store.delete_account(acc_id):
            return jsonify(code=200, result="已删除")
        return jsonify(code=404, result="找不到该抢票人"), 404
    except Exception as e:
        return jsonify(code=500, result=str(e)), 500


# 导入表格时认哪些列名。第一行必须是表头，顺序随意，多余的列忽略。
# 一个字段给多个别名，是因为这份表往往是别人给的，列名五花八门，
# 强行要求"必须叫账号"只会让人反复改表。
IMPORT_COLUMNS = {
    "email": ("账号", "邮箱", "email", "Email", "帐号"),
    "password": ("密码", "password", "Password"),
    "remark": ("备注", "姓名", "名字", "remark", "Remark"),
    "seq": ("窗口序号", "窗口", "序号", "seq"),
    "member_code": ("会员码", "会员优先购票码", "优先购票码", "member_code"),
}

# 解码 CSV 时依次尝试的编码。
# utf-8-sig 放第一位：Excel 存的 UTF-8 CSV 带 BOM，不剥掉第一个列名会变成
# "\ufeff账号" 而匹配不上表头。
# gbk 是**必须有**的一档：中文 Windows 的 Excel「另存为 CSV」默认就是 GBK，
# 不是 UTF-8，按 UTF-8 读会整片乱码，而且乱码之后报的错是"缺少必需的列"，
# 完全看不出真正原因。big5 兜港台环境。
CSV_ENCODINGS = ("utf-8-sig", "utf-8", "gbk", "big5")


def _decode_csv_bytes(raw: bytes):
    """把 CSV 字节按常见编码依次尝试解码，返回 (文本, 用的编码)。"""
    for enc in CSV_ENCODINGS:
        try:
            return raw.decode(enc), enc
        except UnicodeDecodeError:
            continue
    # 全都失败就用 utf-8 忽略错误，至少让用户看到部分内容而不是一个裸异常
    return raw.decode("utf-8", errors="replace"), "utf-8(有无法识别的字符)"


def _rows_from_xlsx(raw: bytes):
    """读 .xlsx 的第一个工作表，返回 (表头列表, 数据行字典列表)。"""
    import io

    from openpyxl import load_workbook

    # read_only + data_only：只读取值，不要公式，也不加载样式，省内存也快
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)

    try:
        header_row = next(rows)
    except StopIteration:
        return [], []

    headers = [("" if h is None else str(h)).strip() for h in header_row]
    out = []
    # 带上工作表里的**真实行号**。跳过空行之后行号会错位，
    # 报"第 3 行有问题"但用户在 Excel 里数到的是第 4 行，会去改错行。
    for sheet_row, r in enumerate(rows, start=2):
        # 整行都空的跳过。Excel 里删过内容的行会留下一堆 None，
        # 不跳过的话导入预览里会冒出几十条"账号为空"的假错误
        if all(c is None or str(c).strip() == "" for c in r):
            continue
        row = {
            headers[i] if i < len(headers) else f"_{i}":
                ("" if c is None else str(c).strip())
            for i, c in enumerate(r)
        }
        row["__line__"] = sheet_row
        out.append(row)
    wb.close()
    return headers, out


@ticket.route("/accounts/import", methods=["post"])
def import_accounts():
    """
    从表格批量导入抢票人。支持 **.csv 和 .xlsx**。

    两种提交方式：
        multipart/form-data，字段 file=<文件>，commit=true/false   （前端用这个）
        application/json，{"csv": "<文本>", "commit": false}       （纯文本，便于脚本调用）

    commit=false（默认）只解析和校验并返回预览，**不写任何数据**；true 才真正写入。

    为什么要两步：直接导入的话，一个格式错误就可能污染整份数据，而且人事后
    不容易发现哪一行错了。先给出"将新增 N 个、更新 M 个、哪几行有问题"，
    看清楚再决定。

    列名见 IMPORT_COLUMNS，第一行必须是表头，顺序随意，多余的列忽略。
    窗口用**序号**而不是 ID：ID 是 32 位十六进制，人没法手填；
    序号就是比特浏览器界面上显示的那个数字。
    """
    import csv
    import io

    try:
        commit = False
        rows_raw = None
        headers = []
        source_note = ""

        upload = request.files.get("file")
        if upload is not None:
            commit = (request.form.get("commit") or "").lower() in ("1", "true", "yes")
            raw = upload.read()
            if not raw:
                return jsonify(code=400, result="文件是空的"), 400
            name = (upload.filename or "").lower()

            if name.endswith((".xlsx", ".xlsm")):
                headers, rows_raw = _rows_from_xlsx(raw)
                source_note = "Excel"
            elif name.endswith(".xls"):
                # 旧的 .xls 是完全不同的二进制格式，openpyxl 读不了，
                # 与其抛一个看不懂的异常，不如直接说清楚该怎么办
                return (
                    jsonify(code=400,
                            result="不支持旧版 .xls 格式，请在 Excel 里另存为 .xlsx 或 .csv"),
                    400,
                )
            else:
                text, enc = _decode_csv_bytes(raw)
                reader = csv.DictReader(io.StringIO(text))
                headers = [(h or "").strip() for h in (reader.fieldnames or [])]
                rows_raw = list(reader)
                source_note = f"CSV（{enc}）"
        else:
            body = request.json or {}
            commit = bool(body.get("commit"))
            text = body.get("csv") or ""
            if not text.strip():
                return jsonify(code=400, result="文件是空的"), 400
            if text.startswith("\ufeff"):
                text = text[1:]
            reader = csv.DictReader(io.StringIO(text))
            headers = [(h or "").strip() for h in (reader.fieldnames or [])]
            rows_raw = list(reader)
            source_note = "CSV"

        if not headers:
            return jsonify(code=400, result="读不出表头，请确认第一行是列名"), 400

        header_map = {h: h for h in headers}

        def col(row, key):
            for n in IMPORT_COLUMNS[key]:
                if n in header_map:
                    v = row.get(header_map[n])
                    return ("" if v is None else str(v)).strip()
            return ""

        # 账号列必须有，但**任何一个别名都算数**。
        # 早先这里写死只认"账号"，而取值时又接受"邮箱"，
        # 结果用「邮箱」当表头的表会被误判成缺列。
        if not any(n in header_map for n in IMPORT_COLUMNS["email"]):
            return (
                jsonify(code=400,
                        result=f"缺少账号列（列名可以是：{'、'.join(IMPORT_COLUMNS['email'][:3])}）。"
                               f"当前表头是：{'、'.join(headers)}"),
                400,
            )

        # 窗口序号 -> ID
        id_by_seq = {}
        conn_err = None
        try:
            for b in getAllBrowsers():
                if b.get("seq") is not None:
                    id_by_seq[str(b["seq"])] = b["id"]
        except Exception as e:
            conn_err = str(e)

        existing = store.get_accounts()
        by_email = {a["email"].lower(): a for a in existing}
        taken = {a.get("browser_id"): a for a in existing if a.get("browser_id")}

        parsed = []
        seen_emails = set()
        seen_seqs = {}
        ok_count = new_count = update_count = 0

        for i, row in enumerate(rows_raw, start=2):   # 第 1 行是表头
            # Excel 路径带了工作表真实行号（跳过空行会让顺序号错位）；
            # CSV 路径没有空行跳过，顺序号就是行号
            i = row.get("__line__", i)
            email = col(row, "email")
            password = col(row, "password")
            remark = col(row, "remark")
            seq = col(row, "seq")
            member_code = col(row, "member_code")
            # Excel 里数字列会读成 "3.0"，转成 "3" 才能和窗口序号对上
            if seq.endswith(".0"):
                seq = seq[:-2]

            issues = []
            if not email:
                issues.append("账号为空")
            elif "@" not in email or "." not in email.split("@")[-1]:
                issues.append("账号不像邮箱")
            elif email.lower() in seen_emails:
                issues.append("文件内账号重复")
            else:
                seen_emails.add(email.lower())

            is_update = email.lower() in by_email
            if not password and not is_update:
                issues.append("密码为空（新增必须填）")

            browser_id = ""
            if seq:
                if conn_err:
                    issues.append("连不上比特浏览器，无法校验窗口序号")
                elif seq not in id_by_seq:
                    issues.append(f"窗口序号 {seq} 不存在")
                else:
                    browser_id = id_by_seq[seq]
                    if seq in seen_seqs:
                        issues.append(f"文件内窗口 {seq} 被分配了多次")
                    else:
                        seen_seqs[seq] = email
                    other = taken.get(browser_id)
                    if other and other["email"].lower() != email.lower():
                        who = other.get("remark") or other["email"]
                        issues.append(f"窗口 {seq} 已绑定给「{who}」")

            if not issues:
                ok_count += 1
                if is_update:
                    update_count += 1
                else:
                    new_count += 1

            parsed.append({
                "line": i, "email": email, "remark": remark, "seq": seq,
                "has_password": bool(password),
                "has_member_code": bool(member_code),
                "action": "更新" if is_update else "新增",
                "issues": issues,
                "_payload": {
                    "email": email, "password": password, "remark": remark,
                    "browser_id": browser_id, "member_code": member_code,
                    "id": by_email[email.lower()]["id"] if is_update else None,
                },
            })

        if not parsed:
            return jsonify(code=400, result="文件里没有数据行"), 400

        summary = f"将新增 {new_count} 人、更新 {update_count} 人"
        bad = len(parsed) - ok_count
        if bad:
            summary += f"；{bad} 行有问题会被跳过"
        if source_note:
            summary += f"（来源：{source_note}）"

        # 预览不回传 _payload，里面有明文密码，没必要过一趟前端
        public = [{k: v for k, v in r.items() if k != "_payload"} for r in parsed]
        if not commit:
            return jsonify(code=200, preview=True, rows=public, summary=summary,
                           ok_count=ok_count, bad_count=bad)

        saved = 0
        errors = []
        for r in parsed:
            if r["issues"]:
                continue
            payload = dict(r["_payload"])
            if payload.get("id") is None:
                payload.pop("id")
            try:
                store.save_account(payload)
                saved += 1
            except ValueError as e:
                errors.append(f"第 {r['line']} 行：{e}")

        msg = f"已导入 {saved} 人"
        if bad:
            msg += f"，跳过 {bad} 行有问题的"
        if errors:
            msg += "；" + "；".join(errors[:3])
        return jsonify(code=200, preview=False, result=msg, saved=saved)
    except Exception as e:
        return jsonify(code=500, result=str(e)), 500


@ticket.route("/accounts/clear_login", methods=["post"])
def clear_login():
    """
    手动清除选中抢票人所绑定窗口的登录态（关窗 + 清 Cookie + 清本地存储）。

    正常换人时 login_all 会按归属记录自动清，这里是人工兜底：
    比如窗口被站点风控盯上想推倒重来、或者你就是想让某个人重新登一次。

    同步执行，不走后台任务：每个窗口就是关窗+两个接口调用，几秒钟的事，
    而且清完必须让前端立刻刷新登录状态，异步反而不好收尾。
    """
    try:
        ids = (request.json or {}).get("ids") or []
        if not ids:
            return jsonify(code=400, result="没有选中任何抢票人"), 400

        accounts = {a["id"]: a for a in store.get_accounts()}
        done, skipped, bids = [], [], []
        for acc_id in ids:
            acc = accounts.get(acc_id)
            if not acc:
                continue
            label = acc.get("remark") or acc.get("email")
            bid = acc.get("browser_id")
            if not bid:
                skipped.append(label)
                continue
            bids.append(bid)
            done.append(label)

        # 批量清：关窗后要等进程退出才能清，这个等待跟窗口数量无关，
        # 一个个清的话 10 个窗口要等 10 次
        if bids:
            clearLoginStateBatch(bids)
            for bid in bids:
                store.clear_window_owner(bid)

        msg = f"已清除 {len(done)} 个窗口的登录态"
        if done:
            msg += f"（{'、'.join(done)}）"
        if skipped:
            msg += f"；{'、'.join(skipped)} 未绑定窗口，已跳过"
        return jsonify(code=200, result=msg)
    except BitBrowserNotRunning as e:
        return jsonify(code=500, result=str(e)), 500
    except Exception as e:
        return jsonify(code=500, result=str(e)), 500


# ------------------------------------------------------------------
# 抢票战果
# ------------------------------------------------------------------

# 状态在界面上的说法。后端存的是英文枚举，给人看要用人话。
STATUS_LABELS = {
    "locked": "待支付",
    "manual": "需人工确认",
    "paid": "已支付",
    "expired": "已过期",
}


@ticket.route("/results", methods=["get"])
def get_results():
    """
    返回全部战果，按批次分组信息一并带出。

    这是整个流程的产出：程序只锁单不付款，每条记录都对应一个等着人去付的订单。
    """
    try:
        data = store.get_results()
        for it in data["items"]:
            it["status_label"] = STATUS_LABELS.get(it.get("status"), it.get("status"))
        return jsonify(code=200, **data)
    except Exception as e:
        return jsonify(code=500, result=str(e)), 500


@ticket.route("/results/status", methods=["post"])
def set_result_status():
    """标记一条战果的状态。付款是人工在浏览器里做的，程序没法自己知道，只能由人来标。"""
    try:
        body = request.json or {}
        result_id = body.get("id")
        status = body.get("status")
        if not result_id or not status:
            return jsonify(code=400, result="缺少 id 或 status"), 400
        if store.set_result_status(result_id, status):
            return jsonify(code=200, result=f"已标记为{STATUS_LABELS.get(status, status)}")
        return jsonify(code=404, result="找不到该记录"), 404
    except ValueError as e:
        return jsonify(code=400, result=str(e)), 400
    except Exception as e:
        return jsonify(code=500, result=str(e)), 500


@ticket.route("/results/batch/delete", methods=["post"])
def delete_result_batch():
    """删掉一个批次的全部战果记录。"""
    try:
        batch_id = (request.json or {}).get("batch_id")
        if not batch_id:
            return jsonify(code=400, result="缺少 batch_id"), 400
        n = store.delete_batch(batch_id)
        return jsonify(code=200, result=f"已删除 {n} 条记录")
    except Exception as e:
        return jsonify(code=500, result=str(e)), 500


@ticket.route("/results/export.csv", methods=["get"])
def export_results_csv():
    """
    导出战果为 CSV。可选 ?batch_id=xxx 只导某一批。

    两个刻意的选择：

    1. **UTF-8 带 BOM**（utf-8-sig）。不带 BOM 的话 Excel 双击打开中文全是乱码，
       而这个文件的使用者多半就是直接双击用 Excel 打开对账的。
    2. **订单号加一个前导制表符**。Excel 会把 16 位纯数字当成数值，
       显示成 1.5804E+15，订单号就废了。加 \t 强制按文本处理。
    """
    import csv
    import io

    try:
        batch_id = request.args.get("batch_id")
        data = store.get_results()
        items = data["items"]
        if batch_id:
            items = [i for i in items if i.get("batch_id") == batch_id]

        batches = data["batches"]
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow([
            "活动", "抢票人", "账号", "窗口序号",
            "场次", "票档", "张数", "订单号", "锁单时间", "状态", "订单页链接",
        ])
        for it in items:
            b = batches.get(it.get("batch_id")) or {}
            order_id = it.get("order_id")
            w.writerow([
                b.get("event_name", ""),
                it.get("account_label", ""),
                it.get("email", ""),
                it.get("browser_seq") if it.get("browser_seq") is not None else "",
                it.get("session_text", ""),
                it.get("tier_text", ""),
                it.get("quantity", ""),
                # 前导 tab：不加的话 Excel 会把它变成科学计数法
                f"\t{order_id}" if order_id else "未抓到，请到窗口内查看",
                (it.get("locked_at") or "").replace("T", " "),
                STATUS_LABELS.get(it.get("status"), it.get("status", "")),
                it.get("page_url", ""),
            ])

        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return Response(
            buf.getvalue().encode("utf-8-sig"),
            # 用 content_type 而不是 mimetype：mimetype 里带 charset 的话
            # Flask 会再追加一次，变成重复的 charset=utf-8; charset=utf-8
            content_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": f'attachment; filename="hkticket_results_{stamp}.csv"'
            },
        )
    except Exception as e:
        return jsonify(code=500, result=str(e)), 500


# ------------------------------------------------------------------
# 第一步：解析活动票务信息
# ------------------------------------------------------------------

@ticket.route("/event", methods=["get"])
def get_event():
    """取回上次解析留存的活动信息，页面刷新后不用重新解析。"""
    try:
        return jsonify(code=200, event=store.get_event())
    except Exception as e:
        return jsonify(code=500, result=str(e)), 500


@ticket.route("/parse_event", methods=["post"])
def do_parse_event():
    """
    解析活动的所有场次和票档并留存。

    纯 HTTP 请求活动数据接口：不开浏览器窗口、不需要登录、不点「立即购买」。
    通常几百毫秒返回，所以直接同步执行，不用后台线程那一套。
    """
    task_id = str(uuid.uuid4())
    try:
        event_url = (request.json or {}).get("event_url", "").strip()

        log_manager.create_task(task_id)
        with LogCapture(task_id):
            event = parse_event(event_url)

        store.save_event(event)
        total = sum(len(s["tiers"]) for s in event.get("sessions", []))
        avail = sum(
            1 for s in event.get("sessions", []) for t in s["tiers"] if t["available"]
        )
        return jsonify(
            code=200,
            result=f"解析成功：{len(event.get('sessions', []))} 个场次 / {total} 个票档（{avail} 个可售）",
            event=event,
            task_id=task_id,
        )
    except ValueError as e:
        return jsonify(code=400, result=str(e), task_id=task_id), 400
    except Exception as e:
        return jsonify(code=500, result=f"解析失败：{str(e)}", task_id=task_id), 500


# ------------------------------------------------------------------
# 第二步：抢票配置
# ------------------------------------------------------------------

@ticket.route("/plans", methods=["get"])
def get_plans():
    try:
        return jsonify(code=200, plans=store.get_plans())
    except Exception as e:
        return jsonify(code=500, result=str(e)), 500


@ticket.route("/plans/save", methods=["post"])
def save_plans():
    """整表覆盖保存每个账号的抢票配置。"""
    try:
        plans = store.save_plans((request.json or {}).get("plans") or {})
        return jsonify(code=200, result=f"已保存 {len(plans)} 条配置")
    except Exception as e:
        return jsonify(code=500, result=str(e)), 500


# ------------------------------------------------------------------
# 一键登录
# ------------------------------------------------------------------

@ticket.route("/login_all", methods=["post"])
def do_login_all():
    """
    一键登录：给每个账号打开各自的窗口，填账号密码并点登录。
    验证码需要人工在弹出的窗口里点掉，本接口会等待（默认 180 秒）。
    """
    task_id = str(uuid.uuid4())
    try:
        body = request.json or {}
        wait_seconds = int(body.get("wait_captcha_seconds", 180))
        only_ids = body.get("account_ids")  # 不传则全部账号

        accounts = store.get_accounts()
        if only_ids:
            accounts = [a for a in accounts if a["id"] in only_ids]
        if not accounts:
            return jsonify(code=400, result="没有可登录的抢票人"), 400

        missing = [a["email"] for a in accounts if not a.get("browser_id")]
        if missing:
            return (
                jsonify(
                    code=400,
                    result=f"这些账号还没绑定浏览器窗口：{'、'.join(missing)}",
                ),
                400,
            )

        _run_in_background(
            task_id,
            lambda pw: login_all(pw, accounts, wait_seconds, task_id),
            "一键登录失败",
            kind="login",
        )
        log_manager.add_log(task_id, f"开始一键登录，共 {len(accounts)} 个账号...", "info")
        return jsonify(code=200, result="一键登录任务已启动", task_id=task_id)
    except Exception as e:
        return jsonify(code=500, result=str(e), task_id=task_id), 500


@ticket.route("/login_status", methods=["get"])
def login_status():
    """
    查所有抢票人的登录状态。

    默认**不拉起**没打开的窗口（10 个人就是弹 10 个窗口，很吵），
    那些标成 window_closed。传 ?open=1 强制全部拉起来查。
    """
    try:
        accounts = store.get_accounts()
        if not accounts:
            return jsonify(code=200, statuses=[])

        open_closed = request.args.get("open") == "1"
        result = {}

        def _thread():
            try:
                async def _run():
                    async with async_playwright() as pw:
                        result["data"] = await check_login_status(
                            pw, accounts, open_closed_windows=open_closed
                        )
                asyncio.run(_run())
            except Exception as e:
                result["error"] = e

        t = threading.Thread(target=_thread, daemon=True)
        t.start()
        t.join(timeout=120)

        if t.is_alive():
            return jsonify(code=500, result="查询登录状态超时"), 500
        if "error" in result:
            return jsonify(code=500, result=f"查询登录状态失败：{result['error']}"), 500
        return jsonify(code=200, statuses=result["data"])
    except BitBrowserNotRunning as e:
        return jsonify(code=503, result=str(e)), 503
    except Exception as e:
        return jsonify(code=500, result=str(e)), 500


@ticket.route("/browsers", methods=["get"])
def browsers():
    """
    列出比特浏览器里所有的窗口（自动翻页，单页上限 100）。
    前端用它来一键填充参与抢票的窗口ID，省去手动复制。
    """
    try:
        rows = getAllBrowsers()
        return jsonify(code=200, result=f"共 {len(rows)} 个窗口", browsers=rows)
    except Exception as e:
        # 最常见的是比特浏览器客户端未登录（token 失效），把原始 msg 透传给前端
        return jsonify(code=500, result=f"获取窗口列表失败：{str(e)}"), 500


def _run_sync(coro_factory, timeout=120):
    """
    在请求线程里同步跑一段 playwright 协程。

    为什么要另起线程：Playwright 的同步/异步 API 不能在已有事件循环里直接跑，
    而 Flask 的请求处理是同步的。起一个线程、在里面 asyncio.run 是最省事的做法。

    :raises TimeoutError: 超时。宁可报超时也不要让请求一直挂着——
        页面上转圈转到天荒地老比报错更难排查。
    """
    box = {}

    def _thread():
        try:
            async def _run():
                async with async_playwright() as pw:
                    box["data"] = await coro_factory(pw)
            asyncio.run(_run())
        except Exception as e:
            box["error"] = e

    t = threading.Thread(target=_thread, daemon=True)
    t.start()
    t.join(timeout=timeout)
    if t.is_alive():
        raise TimeoutError("操作超时")
    if "error" in box:
        raise box["error"]
    return box.get("data")


@ticket.route("/preflight", methods=["get"])
def preflight():
    """
    开抢前自检：逐个抢票人检查"到点能不能真的抢"，给出红绿灯。

    要解决的是最惨的失败模式——**到了开抢那一秒才发现某个人没登录、
    某个人的配置是上个活动的**。抢票循环里确实会检查登录态
    （ticket_operation 第一轮就查），但那时候已经晚了。
    这个接口让你在开抢前十分钟就能看到问题，还来得及修。

    检查项都是**只读且便宜**的，不开新窗口、不点任何购买按钮：
        1. 绑没绑窗口
        2. 窗口在不在（比特浏览器里还存不存在这个 ID）
        3. 登录态（只查已打开的窗口，没开的窗口如实报"窗口未打开"）
        4. 配置完整性 + 是否属于当前解析的活动
        5. 窗口有没有被重复绑定

    返回每个人一条结果，level 取 ok / warn / error：
        error 到点一定抢不了，必须修
        warn  可能有问题（比如窗口没开，登录态查不到）
        ok    没发现问题
    """
    try:
        accounts = store.get_accounts()
        plans = store.get_plans()
        event = store.get_event()

        items = []
        overall_error = 0
        overall_warn = 0

        if not accounts:
            return jsonify(code=200, items=[], summary="还没有添加抢票人",
                           error_count=1, warn_count=0)

        valid_sessions = {s["text"] for s in event.get("sessions", [])}
        valid_tiers = {t["text"] for s in event.get("sessions", []) for t in s["tiers"]}
        member_tiers = {t["text"] for s in event.get("sessions", [])
                        for t in s["tiers"] if t.get("need_member_code")}

        # 窗口是否还存在。比特浏览器连不上时拿不到，这种情况整体报错即可
        existing = None
        conn_err = None
        try:
            existing = {b["id"]: b for b in getAllBrowsers()}
        except Exception as e:
            conn_err = str(e)

        # 已登录状态：复用现成的检查，只查已打开的窗口，不弹新窗口
        status_map = {}
        if conn_err is None:
            try:
                rows = _run_sync(
                    lambda pw: check_login_status(pw, accounts, open_closed_windows=False)
                )
                status_map = {r["account_id"]: r for r in rows}
            except Exception:
                pass

        # 重复绑定检查
        seen = {}
        dup_ids = set()
        for a in accounts:
            bid = a.get("browser_id")
            if not bid:
                continue
            if bid in seen:
                dup_ids.add(a["id"])
                dup_ids.add(seen[bid])
            seen[bid] = a["id"]

        for a in accounts:
            label = a.get("remark") or a.get("email")
            problems = []
            level = "ok"

            bid = a.get("browser_id")
            if not bid:
                problems.append("没绑定浏览器窗口")
                level = "error"
            elif a["id"] in dup_ids:
                problems.append("和别人绑了同一个窗口，并发抢票会互相干扰")
                level = "error"
            elif existing is not None and bid not in existing:
                problems.append("绑定的窗口在比特浏览器里已不存在，请重新选择")
                level = "error"

            plan = plans.get(a["id"]) or {}
            if not event.get("sessions"):
                problems.append("还没解析活动")
                level = "error"
            elif not plan.get("session_text") or not plan.get("tier_text"):
                problems.append("还没配置场次或票档")
                level = "error"
            else:
                if plan["session_text"] not in valid_sessions:
                    problems.append(f"配的场次不属于当前活动（{event.get('name')}）")
                    level = "error"
                if plan["tier_text"] not in valid_tiers:
                    problems.append("配的票档不属于当前活动")
                    level = "error"
                else:
                    sess = next((x for x in event.get("sessions", [])
                                 if x["text"] == plan["session_text"]), None)
                    lim = sess and sess.get("order_limit")
                    if lim and int(plan.get("quantity") or 1) > lim:
                        # 不算错误：撞上限时程序会停下并按实际能买的数量买，
                        # 抢得到，只是买不到那么多张。提前说清楚而已。
                        problems.append(
                            f"配了 {plan.get('quantity')} 张，但本场次限购 {lim} 张，"
                            f"实际只会买到 {lim} 张"
                        )
                        if level != "error":
                            level = "warn"
                if plan["tier_text"] in member_tiers:
                    if (a.get("member_code") or "").strip():
                        problems.append("该票档需要会员优先购票码，已填，弹出时会自动提交")
                    else:
                        # 只是提醒，不是 error：照样可以开抢，
                        # 万一到时候没弹会员码窗（比如正票场次）就正常抢到了
                        problems.append("该票档可能要求会员优先购票码，但这个抢票人没有填")
                        if level != "error":
                            level = "warn"

            st = status_map.get(a["id"], {})
            if conn_err:
                problems.append("连不上比特浏览器，无法检查登录态")
                if level != "error":
                    level = "warn"
            elif st.get("status") == "logged_in":
                # 登录着还不够，还要确认登录的就是本人（换人复用窗口的坑）
                owner = store.get_window_owner(bid)
                if owner and owner.get("email") == (a.get("email") or "").strip():
                    pass
                else:
                    problems.append("窗口登录的可能不是本人，开抢前会自动清除并要求重新登录")
                    if level != "error":
                        level = "warn"
            elif st.get("status") == "logged_out":
                problems.append("未登录，请先执行「一键启动并登录」")
                level = "error"
            elif st.get("status") == "window_closed":
                problems.append("窗口未打开，登录态未知")
                if level != "error":
                    level = "warn"

            if level == "error":
                overall_error += 1
            elif level == "warn":
                overall_warn += 1

            items.append({
                "account_id": a["id"],
                "label": label,
                "email": a.get("email"),
                "level": level,
                "problems": problems,
                "plan": f"{plan.get('session_text', '')} / {plan.get('tier_text', '')} x{plan.get('quantity', 1)}"
                        if plan.get("session_text") else "",
            })

        if overall_error:
            summary = f"{overall_error} 人到点抢不了，需要处理"
        elif overall_warn:
            summary = f"{overall_warn} 人有提示，其余就绪"
        else:
            summary = f"{len(items)} 人全部就绪"

        return jsonify(code=200, items=items, summary=summary,
                       error_count=overall_error, warn_count=overall_warn)
    except Exception as e:
        return jsonify(code=500, result=str(e)), 500


@ticket.route("/active_task", methods=["get"])
def active_task():
    """
    返回当前还在进行中的抢票任务，供页面刷新后接管。

    解决的问题：抢票任务跑在后台线程里，而页面上的 task_id 只是个 JS 变量。
    刷新页面（或电脑睡眠后标签页重载）之后，变量没了但任务还在跑——
    界面回到初始状态，暂停和停止按钮都没了，任务却还在以 2~4 秒轮询，
    用户既看不到进度也停不掉它，只能去杀进程。蹲通宵回流票时这是致命的。

    **以服务端状态为准，不依赖浏览器记住什么。** 这样换个浏览器、
    换台机器打开控制台，照样能接管正在跑的任务。

    同时只会有一个抢票任务（界面上只能启动一个），所以取第一个就够；
    真有多个也一并返回，避免悄悄漏掉。
    """
    try:
        ids = task_state_manager.list_active(kind="grab")
        if not ids:
            return jsonify(code=200, task_id=None, tasks=[])
        tasks = [
            {"task_id": tid, "state": task_state_manager.get_state(tid)}
            for tid in ids
        ]
        return jsonify(code=200, task_id=tasks[0]["task_id"],
                       state=tasks[0]["state"], tasks=tasks)
    except Exception as e:
        return jsonify(code=500, result=str(e)), 500


@ticket.route("/control_task", methods=["post"])
def control_task():
    """控制抢票任务状态：暂停、恢复、停止"""
    try:
        task_id = request.json.get("task_id")
        action = request.json.get("action")  # pause, resume, stop

        if not task_id:
            return jsonify(code=400, result="缺少task_id参数"), 400

        if action == "pause":
            if task_state_manager.pause_task(task_id):
                log_manager.add_log(task_id, "任务已暂停", "warning")
                return jsonify(code=200, result="任务已暂停")
            return jsonify(code=400, result="无法暂停任务（任务可能不存在或未运行）"), 400
        elif action == "resume":
            if task_state_manager.resume_task(task_id):
                log_manager.add_log(task_id, "任务已恢复", "info")
                return jsonify(code=200, result="任务已恢复")
            return jsonify(code=400, result="无法恢复任务（任务可能不存在或未暂停）"), 400
        elif action == "stop":
            if task_state_manager.stop_task(task_id):
                log_manager.add_log(task_id, "任务已停止", "warning")
                return jsonify(code=200, result="任务已停止")
            return jsonify(code=400, result="无法停止任务（任务可能不存在）"), 400
        else:
            return (
                jsonify(code=400, result="无效的action参数，支持：pause, resume, stop"),
                400,
            )
    except Exception as e:
        return jsonify(code=500, result=str(e)), 500


@ticket.route("/get_logs", methods=["get"])
def get_logs():
    """获取指定任务的日志（轮询方式，SSE 不可用时的兜底）"""
    task_id = request.args.get("task_id")
    since_index = int(request.args.get("since_index", 0))

    if not task_id:
        return jsonify(code=400, result="缺少task_id参数"), 400

    logs_data = log_manager.get_logs(task_id, since_index)
    return jsonify(code=200, **logs_data)


@ticket.route("/logs_stream", methods=["get"])
def logs_stream():
    """SSE流式推送日志
    如果task_id为空，则推送所有任务的日志（全局模式）
    """
    task_id = request.args.get("task_id")  # 如果为空，则为全局模式

    def generate():
        """生成SSE事件流"""
        # 注册SSE客户端（task_id为None时注册为全局客户端）
        log_queue = log_manager.register_sse_client(task_id if task_id else None)

        try:
            # 发送初始连接成功消息
            if task_id:
                yield f"data: {json.dumps({'type': 'connected', 'message': f'SSE连接已建立（任务: {task_id}）'})}\n\n"
            else:
                yield f"data: {json.dumps({'type': 'connected', 'message': 'SSE连接已建立（全局模式）'})}\n\n"

            # 如果指定了task_id，发送历史日志
            if task_id:
                history_logs = log_manager.get_logs(task_id, since_index=0)
                if history_logs.get("logs"):
                    recent_logs = history_logs["logs"][-50:]  # 只发送最近50条
                    for log_entry in recent_logs:
                        # 为历史日志添加task_id，以便前端过滤
                        log_entry_with_task = {**log_entry, "task_id": task_id}
                        yield f"data: {json.dumps({'type': 'log', 'data': log_entry_with_task})}\n\n"

            # 持续监听新日志
            import time as time_module

            last_activity_time = time_module.time()
            max_idle_time = 300  # 5分钟无日志后断开连接

            while True:
                try:
                    # 使用超时机制，每30秒发送一次心跳
                    log_entry = log_queue.get(timeout=30)

                    if log_entry is None:
                        # None是结束信号
                        yield f"data: {json.dumps({'type': 'end', 'message': '日志流已结束'})}\n\n"
                        break

                    # 发送日志
                    yield f"data: {json.dumps({'type': 'log', 'data': log_entry})}\n\n"
                    last_activity_time = time_module.time()  # 更新活动时间

                except queue.Empty:
                    # 超时，发送心跳
                    yield f"data: {json.dumps({'type': 'heartbeat', 'message': 'keepalive'})}\n\n"

                    # 检查是否超过最大空闲时间
                    if time_module.time() - last_activity_time > max_idle_time:
                        # 长时间无日志，断开连接
                        yield f"data: {json.dumps({'type': 'timeout', 'message': '连接超时，已断开'})}\n\n"
                        break

        except GeneratorExit:
            # 客户端断开连接
            pass
        except Exception as e:
            # 发生错误
            yield f"data: {json.dumps({'type': 'error', 'message': str(e)})}\n\n"
        finally:
            # 注销SSE客户端
            log_manager.unregister_sse_client(task_id if task_id else None, log_queue)

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",  # 禁用Nginx缓冲
            "Connection": "keep-alive",
        },
    )
